import sys
import os
import re
import openpyxl
import pandas as pd
from PIL import Image, ImageDraw, ImageFont
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QLabel, QLineEdit, QPushButton, QFileDialog, QProgressBar,
                             QSpinBox, QColorDialog, QMessageBox, QSlider, QGroupBox, QCheckBox)
from PyQt5.QtCore import Qt, QSize, QThread, pyqtSignal
from PyQt5.QtGui import QPixmap, QIcon
import json
import cv2
import numpy as np
import subprocess
import tempfile
import shutil
import traceback

# Excel处理函数（保持不变）
def is_empty_row(row):
    """检查一行是否为空行（所有单元格都为空或只包含空白字符）"""
    for cell in row:
        if cell.value is not None:
            if isinstance(cell.value, str):
                if cell.value.strip():
                    return False
            else:
                return False
    return True

def remove_empty_rows(ws):
    """从工作表中删除所有空行"""
    empty_row_indices = []
    for row_idx in range(ws.max_row, 0, -1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))[0]
        if is_empty_row(row):
            empty_row_indices.append(row_idx)

    for row_idx in empty_row_indices:
        ws.delete_rows(row_idx)

    print(f"删除了 {len(empty_row_indices)} 个空行")
    return len(empty_row_indices)

def remove_zero_width_spaces(file_path):
    """移除零宽度空格并删除空行，直接覆盖原文件"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    zero_width_chars = [
        '\u200b', '\u200c', '\u200d', '\ufeff',
        '\u200e', '\u200f', '\u202a', '\u202b',
        '\u202c', '\u202d', '\u202e',
    ]

    cleaned_cells = 0

    for row in ws.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                original_value = cell.value
                cleaned_value = original_value

                for char in zero_width_chars:
                    cleaned_value = cleaned_value.replace(char, '')

                cleaned_value = re.sub(r'\s+', '', cleaned_value)

                if original_value != cleaned_value:
                    cell.value = cleaned_value
                    cleaned_cells += 1

    print(f"清理了 {cleaned_cells} 个单元格中的空格")

    removed_rows = remove_empty_rows(ws)
    wb.save(file_path)
    print(f"已完成处理并覆盖原文件: {file_path}")

    wb_new = openpyxl.load_workbook(file_path)
    ws_new = wb_new.active
    print(f"\n处理后的表格信息：")
    print(f"总行数: {ws_new.max_row}")
    print(f"总列数: {ws_new.max_column}")

    print("\n最终清理后的前5行内容预览：")
    for i, row in enumerate(ws_new.iter_rows(min_col=1, max_col=1, max_row=min(5, ws_new.max_row)), 1):
        for cell in row:
            cell_value = cell.value if cell.value is not None else "[空]"
            print(f"第{i}行: '{cell_value}'")

def detect_special_chars(file_path):
    """检测特殊字符（用于调试）"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    print("检测特殊字符：")
    special_chars_found = False

    for i, row in enumerate(ws.iter_rows(min_col=1, max_col=1, max_row=min(10, ws.max_row)), 1):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for char in ['\u200b', '\u200c', '\u200d']:
                    if char in cell.value:
                        print(f"第{i}行发现零宽度字符: {repr(char)}")
                        special_chars_found = True

                if any(char in cell.value for char in ['\u200b', '\u200c', '\u200d']):
                    print(f"第{i}行原始内容: {repr(cell.value)}")

    if not special_chars_found:
        print("未发现零宽度字符")

def backup_original_file(file_path):
    """备份原文件"""
    backup_path = file_path.replace('.xlsx', '_backup.xlsx')
    if os.path.exists(backup_path):
        os.remove(backup_path)
    os.rename(file_path, backup_path)
    return backup_path

def restore_from_backup(backup_path, original_path):
    """从备份恢复文件"""
    if os.path.exists(original_path):
        os.remove(original_path)
    os.rename(backup_path, original_path)

# 视频生成线程类（修复封面和正文显示问题）
class VideoGeneratorThread(QThread):
    progress_signal = pyqtSignal(int)
    finished_signal = pyqtSignal(str)
    error_signal = pyqtSignal(str)

    def __init__(self, generator, texts, bg_images, output_type, bgm_path=None, add_cover=True, cover_font_size=60):
        super().__init__()
        self.generator = generator
        self.texts = texts
        self.bg_images = bg_images
        self.output_type = output_type
        self.bgm_path = bgm_path
        self.add_cover = add_cover
        self.cover_font_size = cover_font_size
        self.fps = 30
        self.char_duration = 0.1
        self.final_duration = 2
        self.transition_duration = 1
        self.cover_duration = 2  # 封面显示2秒

    def run(self):
        try:
            if self.output_type == "video":
                video_without_audio = 'output/temp_video_no_audio.mp4'
                self.generate_combined_video(video_without_audio)

                final_output = 'output/final_video_wechat.mp4'
                if self.optimize_for_wechat(video_without_audio, final_output):
                    if self.bgm_path and os.path.exists(self.bgm_path):
                        final_with_bgm = 'output/final_video_with_bgm.mp4'
                        if self.add_background_music(final_output, final_with_bgm):
                            self.finished_signal.emit(f'视频生成完成！视频号兼容版本：{final_with_bgm}')
                        else:
                            self.finished_signal.emit(f'视频生成完成！视频号兼容版本：{final_output}')
                    else:
                        self.finished_signal.emit(f'视频生成完成！视频号兼容版本：{final_output}')
                else:
                    shutil.copy(video_without_audio, final_output)
                    self.finished_signal.emit(f'视频生成完成！保存为：{final_output}')
            else:
                total = len(self.texts)
                for i, (text, bg_path) in enumerate(zip(self.texts, self.bg_images)):
                    output_path = f'output/poster_{i+1}.png'
                    img = self.generator.generate_poster(text, bg_path)
                    img.save(output_path)
                    self.progress_signal.emit(int((i + 1) / total * 100))

                self.finished_signal.emit(f'海报生成完成！共 {len(self.texts)} 张')

        except Exception as e:
            self.error_signal.emit(f'生成失败: {str(e)}')
            print(traceback.format_exc())
        finally:
            if os.path.exists('output/temp_video_no_audio.mp4'):
                os.remove('output/temp_video_no_audio.mp4')

    def generate_combined_video(self, output_path):
        all_frames = []
        bg_images = self.bg_images if self.bg_images else [None] * len(self.texts)

        # 修复：确保bg_images长度与texts一致
        if len(bg_images) < len(self.texts):
            bg_images = (bg_images * (len(self.texts) // len(bg_images) + 1))[:len(self.texts)]

        total_sections = len(self.texts) + (1 if self.add_cover and self.texts else 0)
        current_section = 0

        # 添加封面帧（如果有文本且启用了封面功能）
        if self.add_cover and self.texts:
            cover_text = self.texts[0]  # 使用第一行Excel数据作为封面
            cover_bg = bg_images[0] if bg_images else None

            # 生成封面海报
            cover_img = self.generator.generate_cover_poster(cover_text, cover_bg, self.cover_font_size)
            cover_frame = cv2.cvtColor(np.array(cover_img.convert('RGB')), cv2.COLOR_RGB2BGR)

            # 添加封面帧（显示2秒）
            cover_frames = [cover_frame] * int(self.fps * self.cover_duration)
            all_frames.extend(cover_frames)

            current_section += 1
            progress = int(current_section / total_sections * 100)
            self.progress_signal.emit(progress)

        # 生成正文视频内容（修复：显示所有文本内容，包括第一行）
        content_texts = self.texts  # 修复：显示所有文本，包括第一行
        content_bg_images = bg_images

        for i, (text, bg_path) in enumerate(zip(content_texts, content_bg_images)):
            # 如果是第一行且已经添加了封面，跳过封面显示部分，直接生成打字效果
            if i == 0 and self.add_cover and self.texts:
                # 为第一行内容生成打字效果视频（不跳过）
                frames = self.generator.generate_typing_frames(
                    text, bg_path, self.fps, self.char_duration, self.final_duration
                )
            else:
                # 为其他行内容生成打字效果视频
                frames = self.generator.generate_typing_frames(
                    text, bg_path, self.fps, self.char_duration, self.final_duration
                )

            # 添加转场效果（如果不是第一帧）
            if len(all_frames) > 0:
                transition_frames = self.create_transition(
                    all_frames[-1], frames[0], self.fps, self.transition_duration
                )
                all_frames.extend(transition_frames)

            all_frames.extend(frames)

            current_section += 1
            progress = int(current_section / total_sections * 100)
            self.progress_signal.emit(progress)

        self.write_video_frames(output_path, all_frames)

    def create_transition(self, frame1, frame2, fps, duration):
        transition_frames = []
        steps = int(fps * duration)

        for i in range(steps):
            alpha = i / steps
            blended = cv2.addWeighted(frame1, 1 - alpha, frame2, alpha, 0)
            transition_frames.append(blended)

        return transition_frames

    def write_video_frames(self, output_path, frames):
        if not frames:
            return False

        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        height, width = frames[0].shape[:2]

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        if os.path.exists(output_path):
            os.remove(output_path)

        video_writer = cv2.VideoWriter(output_path, fourcc, self.fps, (width, height))

        for frame in frames:
            video_writer.write(frame)

        video_writer.release()
        return True

    def optimize_for_wechat(self, input_path, output_path):
        try:
            temp_dir = tempfile.mkdtemp()
            temp_output = os.path.join(temp_dir, 'temp_wechat.mp4')

            cmd = [
                'ffmpeg',
                '-y',
                '-i', input_path,
                '-c:v', 'libx264',
                '-preset', 'medium',
                '-crf', '23',
                '-pix_fmt', 'yuv420p',
                '-movflags', '+faststart',
                '-loglevel', 'warning',
                temp_output
            ]

            result = subprocess.run(cmd, capture_output=True, text=True)

            if result.returncode == 0:
                shutil.move(temp_output, output_path)
                print(f"视频号优化成功：{output_path}")
                return True
            else:
                print(f"视频号优化失败：{result.stderr}")
                return False

        except Exception as e:
            print(f"视频号优化异常：{str(e)}")
            return False
        finally:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)

    def add_background_music(self, video_path, output_path):
        try:
            temp_dir = tempfile.mkdtemp()
            temp_output = os.path.join(temp_dir, 'temp_with_bgm.mp4')

            cmd = [
                'ffmpeg',
                '-y',
                '-i', video_path,
                '-i', self.bgm_path,
                '-c:v', 'copy',
                '-c:a', 'aac',
                '-b:a', '192k',
                '-shortest',
                '-map', '0:v:0',
                '-map', '1:a:0',
                '-movflags', '+faststart',
                '-loglevel', 'warning',
                temp_output
            ]

            result = subprocess.run(cmd, capture_output=True, text=True)

            if result.returncode == 0:
                shutil.move(temp_output, output_path)
                return True
            else:
                print(f"音频合并失败：{result.stderr}")
                return False

        except Exception as e:
            print(f"音频合并异常：{str(e)}")
            return False
        finally:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)

# 主窗口类（保持不变）
class VideoGenerator(QMainWindow):
    def __init__(self):
        super().__init__()
        self.bg_images = []
        self.bgm_path = None
        self.current_texts = []  # 保存当前加载的文本内容
        self.initUI()
        self.load_last_config()

    def initUI(self):
        self.setWindowTitle('书单视频生成器（打字机效果+文字上移版+封面功能）')
        self.setGeometry(300, 300, 800, 800)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout()

        # 文件设置区域
        file_group = QGroupBox("文件设置")
        file_layout = QVBoxLayout()

        # Excel文件选择
        excel_layout = QHBoxLayout()
        excel_layout.addWidget(QLabel('Excel文件:'))
        self.excel_path = QLineEdit()
        self.excel_btn = QPushButton('选择文件')
        self.excel_btn.clicked.connect(self.select_excel)
        self.reload_excel_btn = QPushButton('重新加载')
        self.reload_excel_btn.clicked.connect(self.reload_excel_data)
        excel_layout.addWidget(self.excel_path)
        excel_layout.addWidget(self.excel_btn)
        excel_layout.addWidget(self.reload_excel_btn)

        # 背景图片选择
        bg_layout = QHBoxLayout()
        bg_layout.addWidget(QLabel('背景图片:'))
        self.bg_paths = QLineEdit()
        self.bg_btn = QPushButton('选择图片')
        self.bg_btn.clicked.connect(self.select_bg_images)
        self.bg_clear_btn = QPushButton('清空')
        self.bg_clear_btn.clicked.connect(self.clear_bg_images)
        bg_layout.addWidget(self.bg_paths)
        bg_layout.addWidget(self.bg_btn)
        bg_layout.addWidget(self.bg_clear_btn)

        # 背景音乐选择
        self.bgm_layout = QHBoxLayout()
        self.bgm_layout.addWidget(QLabel('背景音乐:'))
        self.bgm_path_label = QLineEdit()
        self.bgm_path_label.setPlaceholderText('视频模式下可选')
        self.bgm_btn = QPushButton('选择音乐')
        self.bgm_btn.clicked.connect(self.select_bgm)
        self.bgm_clear_btn = QPushButton('清空')
        self.bgm_clear_btn.clicked.connect(self.clear_bgm)
        self.bgm_layout.addWidget(self.bgm_path_label)
        self.bgm_layout.addWidget(self.bgm_btn)
        self.bgm_layout.addWidget(self.bgm_clear_btn)

        # 预览区域
        self.bg_preview = QLabel()
        self.bg_preview.setFixedSize(200, 100)
        self.bg_preview.setStyleSheet("border: 1px solid gray;")

        file_layout.addLayout(excel_layout)
        file_layout.addLayout(bg_layout)
        file_layout.addLayout(self.bgm_layout)
        file_layout.addWidget(QLabel("已选背景图预览:"))
        file_layout.addWidget(self.bg_preview)
        file_group.setLayout(file_layout)

        # 封面设置区域（新增）
        cover_group = QGroupBox("封面设置")
        cover_layout = QVBoxLayout()

        cover_option_layout = QHBoxLayout()
        self.add_cover_checkbox = QCheckBox("添加封面（使用第一行Excel数据）")
        self.add_cover_checkbox.setChecked(True)
        cover_option_layout.addWidget(self.add_cover_checkbox)

        cover_font_layout = QHBoxLayout()
        cover_font_layout.addWidget(QLabel('封面字体大小:'))
        self.cover_font_size = QSpinBox()
        self.cover_font_size.setRange(20, 150)
        self.cover_font_size.setValue(60)
        cover_font_layout.addWidget(self.cover_font_size)

        cover_font_layout.addWidget(QLabel('每行字数:'))
        self.chars_per_line = QSpinBox()
        self.chars_per_line.setRange(1, 20)
        self.chars_per_line.setValue(5)
        cover_font_layout.addWidget(self.chars_per_line)

        cover_layout.addLayout(cover_option_layout)
        cover_layout.addLayout(cover_font_layout)
        cover_group.setLayout(cover_layout)

        # 文字设置区域
        text_group = QGroupBox("正文文字设置")
        text_layout = QVBoxLayout()

        text_layout.addWidget(QLabel('顶部文字:'))
        self.header_text = QLineEdit('精选书单推荐')
        text_layout.addWidget(self.header_text)

        text_layout.addWidget(QLabel('底部文字:'))
        self.footer_text = QLineEdit('作者: 某某某 | 《好书推荐》')
        text_layout.addWidget(self.footer_text)

        text_group.setLayout(text_layout)

        # 样式设置区域
        style_group = QGroupBox("正文样式设置")
        style_layout = QVBoxLayout()

        # 字体大小设置
        size_layout = QHBoxLayout()
        size_layout.addWidget(QLabel('顶部文字大小:'))
        self.header_size = QSpinBox()
        self.header_size.setRange(10, 100)
        self.header_size.setValue(30)
        size_layout.addWidget(self.header_size)

        size_layout.addWidget(QLabel('主文字大小:'))
        self.main_size = QSpinBox()
        self.main_size.setRange(20, 150)
        self.main_size.setValue(80)
        size_layout.addWidget(self.main_size)

        size_layout.addWidget(QLabel('底部文字大小:'))
        self.footer_size = QSpinBox()
        self.footer_size.setRange(10, 100)
        self.footer_size.setValue(30)
        size_layout.addWidget(self.footer_size)

        # 颜色设置
        color_layout = QHBoxLayout()
        color_layout.addWidget(QLabel('文字颜色:'))
        self.text_color_btn = QPushButton('选择')
        self.text_color_btn.clicked.connect(lambda: self.choose_color('text'))
        color_layout.addWidget(self.text_color_btn)

        color_layout.addWidget(QLabel('主文字颜色:'))
        self.main_color_btn = QPushButton('选择')
        self.main_color_btn.clicked.connect(lambda: self.choose_color('main'))
        color_layout.addWidget(self.main_color_btn)

        # 背景色设置
        bg_color_layout = QHBoxLayout()
        bg_color_layout.addWidget(QLabel('背景颜色:'))
        self.bg_color_btn = QPushButton('选择')
        self.bg_color_btn.clicked.connect(lambda: self.choose_color('bg'))
        bg_color_layout.addWidget(self.bg_color_btn)

        bg_color_layout.addWidget(QLabel('透明度:'))
        self.bg_alpha = QSlider(Qt.Horizontal)
        self.bg_alpha.setRange(0, 100)
        self.bg_alpha.setValue(50)
        self.bg_alpha.valueChanged.connect(self.update_alpha_label)
        self.alpha_label = QLabel('50%')
        bg_color_layout.addWidget(self.bg_alpha)
        bg_color_layout.addWidget(self.alpha_label)

        # 视频设置
        video_layout = QHBoxLayout()
        video_layout.addWidget(QLabel('输出类型:'))
        self.output_type = QPushButton('海报')
        self.output_type.setCheckable(True)
        self.output_type.toggled.connect(self.toggle_output_type)
        video_layout.addWidget(self.output_type)

        video_layout.addWidget(QLabel('帧率:'))
        self.fps_spin = QSpinBox()
        self.fps_spin.setRange(10, 60)
        self.fps_spin.setValue(30)
        video_layout.addWidget(self.fps_spin)

        video_layout.addWidget(QLabel('字符速度:'))
        self.char_speed = QSlider(Qt.Horizontal)
        self.char_speed.setRange(1, 20)
        self.char_speed.setValue(10)
        self.char_speed.valueChanged.connect(self.update_speed_label)
        self.speed_label = QLabel('0.10s/字符')
        video_layout.addWidget(self.char_speed)
        video_layout.addWidget(self.speed_label)

        style_layout.addLayout(size_layout)
        style_layout.addLayout(color_layout)
        style_layout.addLayout(bg_color_layout)
        style_layout.addLayout(video_layout)
        style_group.setLayout(style_layout)

        # 进度条
        self.progress = QProgressBar()

        # 生成按钮
        self.generate_btn = QPushButton('生成视频（打字机效果+文字上移+封面）')
        self.generate_btn.clicked.connect(self.generate_videos)

        # 状态标签
        self.status_label = QLabel('准备就绪 - 已修复打字机效果，中间文字上移100像素，支持封面功能')

        # 添加到主布局
        layout.addWidget(file_group)
        layout.addWidget(cover_group)
        layout.addWidget(text_group)
        layout.addWidget(style_group)
        layout.addWidget(self.progress)
        layout.addWidget(self.generate_btn)
        layout.addWidget(self.status_label)

        central_widget.setLayout(layout)

        # 初始化颜色
        self.colors = {
            'bg': (255, 255, 255, 128),
            'text': (0, 0, 0, 255),
            'main': (0, 0, 0, 255)
        }

        # 创建输出目录
        os.makedirs('output', exist_ok=True)
        os.makedirs('fonts', exist_ok=True)
        os.makedirs('config', exist_ok=True)

        # 初始隐藏背景音乐选项
        self.toggle_bgm_visibility(False)

    def toggle_output_type(self, checked):
        self.output_type.setText('视频' if checked else '海报')
        self.toggle_bgm_visibility(checked)

    def toggle_bgm_visibility(self, visible):
        for i in range(self.bgm_layout.count()):
            widget = self.bgm_layout.itemAt(i).widget()
            if widget:
                widget.setVisible(visible)

    def update_speed_label(self, value):
        char_duration = 0.1 * (20 / value)
        self.speed_label.setText(f"{char_duration:.2f}s/字符")

    def update_alpha_label(self, value):
        self.alpha_label.setText(f"{value}%")
        alpha = int(value * 255 / 100)
        self.colors['bg'] = self.colors['bg'][:3] + (alpha,)

    def select_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, '选择Excel文件', '', 'Excel Files (*.xlsx *.xls)')
        if path:
            self.excel_path.setText(path)
            self.reload_excel_data()

    def select_bg_images(self):
        paths, _ = QFileDialog.getOpenFileNames(self, '选择背景图片', '', 'Image Files (*.png *.jpg *.jpeg)')
        if paths:
            self.bg_images = paths
            self.bg_paths.setText(f"已选择 {len(paths)} 张图片")
            self.update_bg_preview()

    def clear_bg_images(self):
        self.bg_images = []
        self.bg_paths.clear()
        self.bg_preview.clear()

    def select_bgm(self):
        path, _ = QFileDialog.getOpenFileName(self, '选择背景音乐', '', '音频文件 (*.mp3 *.wav *.aac *.m4a)')
        if path:
            self.bgm_path = path
            self.bgm_path_label.setText(os.path.basename(path))

    def clear_bgm(self):
        self.bgm_path = None
        self.bgm_path_label.clear()

    def update_bg_preview(self):
        if self.bg_images:
            pixmap = QPixmap(self.bg_images[0])
            if not pixmap.isNull():
                pixmap = pixmap.scaled(200, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.bg_preview.setPixmap(pixmap)

    def choose_color(self, type_):
        color = QColorDialog.getColor()
        if color.isValid():
            if type_ == 'bg':
                alpha = self.colors['bg'][3]
                self.colors[type_] = (color.red(), color.green(), color.blue(), alpha)
            else:
                self.colors[type_] = (color.red(), color.green(), color.blue(), 255)

    def load_last_config(self):
        try:
            if os.path.exists('config/last_config.json'):
                with open('config/last_config.json', 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.header_text.setText(config.get('header_text', '精选书单推荐'))
                    self.footer_text.setText(config.get('footer_text', '作者: 某某某 | 《好书推荐》'))
                    self.header_size.setValue(config.get('header_size', 30))
                    self.main_size.setValue(config.get('main_size', 80))
                    self.footer_size.setValue(config.get('footer_size', 30))
                    self.bg_alpha.setValue(config.get('bg_alpha', 50))
                    self.output_type.setChecked(config.get('output_type', False))
                    self.fps_spin.setValue(config.get('fps', 30))
                    self.char_speed.setValue(config.get('char_speed', 10))
                    self.add_cover_checkbox.setChecked(config.get('add_cover', True))
                    self.cover_font_size.setValue(config.get('cover_font_size', 60))
                    self.chars_per_line.setValue(config.get('chars_per_line', 5))
        except:
            pass

    def save_config(self):
        config = {
            'header_text': self.header_text.text(),
            'footer_text': self.footer_text.text(),
            'header_size': self.header_size.value(),
            'main_size': self.main_size.value(),
            'footer_size': self.footer_size.value(),
            'bg_alpha': self.bg_alpha.value(),
            'output_type': self.output_type.isChecked(),
            'fps': self.fps_spin.value(),
            'char_speed': self.char_speed.value(),
            'add_cover': self.add_cover_checkbox.isChecked(),
            'cover_font_size': self.cover_font_size.value(),
            'chars_per_line': self.chars_per_line.value()
        }
        with open('config/last_config.json', 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False)

    def reload_excel_data(self):
        """显式重新加载Excel数据"""
        if not self.excel_path.text():
            QMessageBox.warning(self, '警告', '请先选择Excel文件')
            return False

        try:
            # 读取Excel数据
            df = pd.read_excel(self.excel_path.text(), header=None)
            texts = df.iloc[1:, 0].tolist()
            texts = [str(t).replace(' ', '').strip() for t in texts if pd.notna(t) and str(t).strip() != '']

            if not texts:
                QMessageBox.warning(self, '警告', 'Excel中没有有效文本内容')
                return False

            self.current_texts = texts
            QMessageBox.information(self, '成功', f'已重新加载 {len(texts)} 条文本内容')
            return True
        except Exception as e:
            QMessageBox.critical(self, '错误', f'读取Excel失败: {str(e)}')
            return False

    def generate_videos(self):
        if not self.excel_path.text():
            QMessageBox.warning(self, '警告', '请先选择Excel文件')
            return

        try:
            self.generate_btn.setEnabled(False)
            self.status_label.setText('正在生成视频（打字机效果+文字上移+封面）...')
            self.progress.setValue(0)

            # 使用内存中的文本内容
            texts = self.current_texts if self.current_texts else []
            if not texts:
                QMessageBox.warning(self, '警告', '没有可用的文本内容，请先加载Excel文件')
                return

            # 处理背景图片
            bg_images = self.bg_images
            if not bg_images:
                bg_images = [None] * len(texts)
            elif len(bg_images) < len(texts):
                bg_images = (bg_images * (len(texts) // len(bg_images) + 1))[:len(texts)]

            # 创建并启动生成线程
            output_type = "video" if self.output_type.isChecked() else "poster"
            self.generator_thread = VideoGeneratorThread(
                self, texts, bg_images, output_type, self.bgm_path,
                self.add_cover_checkbox.isChecked(), self.cover_font_size.value()
            )
            self.generator_thread.progress_signal.connect(self.progress.setValue)
            self.generator_thread.finished_signal.connect(self.on_generation_finished)
            self.generator_thread.error_signal.connect(self.on_generation_error)
            self.generator_thread.start()

        except Exception as e:
            QMessageBox.critical(self, '错误', f'生成失败: {str(e)}')
            self.status_label.setText('生成失败')
            self.generate_btn.setEnabled(True)

    def on_generation_finished(self, message):
        QMessageBox.information(self, '完成', message)
        self.status_label.setText(message)
        self.generate_btn.setEnabled(True)
        self.save_config()

    def on_generation_error(self, message):
        QMessageBox.critical(self, '错误', message)
        self.status_label.setText('生成失败')
        self.generate_btn.setEnabled(True)

    def generate_cover_poster(self, text, bg_path=None, font_size=60):
        """生成封面海报（修改版：全黑背景 + 等比例缩放图片）"""
        # 创建全黑色背景画布
        base = Image.new('RGB', (900, 1600), (0, 0, 0))
        draw = ImageDraw.Draw(base)

        try:
            cover_font = ImageFont.truetype('fonts/simhei.ttf', font_size)
        except:
            cover_font = ImageFont.load_default()

        # 封面文字自动换行（每行最多chars_per_line个字）
        words = list(text)
        lines = []
        current_line = []
        max_chars_per_line = self.chars_per_line.value()

        for word in words:
            current_line.append(word)
            if len(current_line) >= max_chars_per_line:
                lines.append(''.join(current_line))
                current_line = []

        if current_line:
            lines.append(''.join(current_line))

        # 计算文字总高度，垂直居中（上半部分）
        total_text_height = len(lines) * font_size * 1.2
        text_start_y = (800 - total_text_height) / 2+100  # 上半部分居中

        # 绘制封面文字（白色文字）
        for i, line in enumerate(lines):
            y_position = text_start_y + i * font_size * 1.2
            draw.text((450, y_position), line, font=cover_font,
                     fill=(255, 255, 255), anchor='mm')  # 白色文字

        # 在下方添加背景图预览（如果有背景图）
        if bg_path and self.bg_images:
            # 使用第一张背景图作为预览
            preview_bg_path = self.bg_images[0]
            preview_img = Image.open(preview_bg_path).convert('RGB')

            # 设置预览图宽度为画布宽度的2/3（等比例缩放）
            target_width = int(900 * 2 / 3)  # 600像素
            original_width, original_height = preview_img.size

            # 等比例计算高度
            scale_factor = target_width / original_width
            target_height = int(original_height * scale_factor)

            # 等比例缩放图片（保持原比例，不拉伸）
            preview_img = preview_img.resize((target_width, target_height), Image.LANCZOS)

            # 计算预览图位置（居中于下半部分）
            preview_x = (900 - target_width) // 2
            preview_y = 800 + (400 - target_height) // 2  # 下半部分居中

            # 如果图片高度超过下半部分，从顶部开始显示
            if target_height > 800:
                preview_y = 800

            # 添加白色边框
            border_size = 5
            border_img = Image.new('RGB', (target_width + border_size*2, target_height + border_size*2), (255, 255, 255))
            base.paste(border_img, (preview_x - border_size, preview_y - border_size))

            # 粘贴预览图
            base.paste(preview_img, (preview_x, preview_y))

        return base

    def generate_poster(self, text, bg_path=None):
        """生成正文海报"""
        if bg_path:
            bg = Image.open(bg_path).convert('RGBA')
            target_width, target_height = 900, 1600
            bg_width, bg_height = bg.size

            width_ratio = target_width / bg_width
            height_ratio = target_height / bg_height
            scale_ratio = min(width_ratio, height_ratio)

            new_width = int(bg_width * scale_ratio)
            new_height = int(bg_height * scale_ratio)
            bg = bg.resize((new_width, new_height), Image.LANCZOS)

            new_bg = Image.new('RGBA', (target_width, target_height))
            offset = ((target_width - new_width) // 2, (target_height - new_height) // 2)
            new_bg.paste(bg, offset)
            bg = new_bg
        else:
            bg = Image.new('RGBA', (900, 1600))

        bg_color = Image.new('RGBA', (900, 1600), self.colors['bg'])
        base = Image.alpha_composite(bg, bg_color)
        draw = ImageDraw.Draw(base)

        try:
            header_font = ImageFont.truetype('fonts/simhei.ttf', self.header_size.value())
            main_font = ImageFont.truetype('fonts/simhei.ttf', self.main_size.value())
            footer_font = ImageFont.truetype('fonts/simhei.ttf', self.footer_size.value())
        except:
            header_font = ImageFont.load_default()
            main_font = ImageFont.load_default()
            footer_font = ImageFont.load_default()

        # 顶部文字
        draw.text((450, 160), self.header_text.text(), font=header_font,
                 fill=self.colors['text'], anchor='mm')

        break_after_punctuation = {'，', '。', '；', '！', '？', '」', '》', '》', '、', ',', '.', ';', '!', '?'}
        # 中间文字（自动换行）- 位置上移100像素
        words = list(text)
        lines = []
        current_line = []

        for word in words:
            current_line.append(word)
            if (len(current_line) >= 10 or
                (word in break_after_punctuation and len(current_line) > 1)):
                lines.append(''.join(current_line))
                current_line = []

        if current_line:
            lines.append(''.join(current_line))

        y_position = 500  # 修改：从600改为500，上移100像素
        for line in lines:
            draw.text((450, y_position), line, font=main_font,
                     fill=self.colors['main'], anchor='mm')
            y_position += 100

        # 底部文字
        draw.text((450, 1440), self.footer_text.text(), font=footer_font,
                 fill=self.colors['text'], anchor='mm')

        return base.convert('RGB')

    def generate_typing_frames(self, text, bg_path=None, fps=30, char_duration=0.1, final_duration=2):
        """生成真正的打字机效果帧序列（每个字符逐个出现）"""
        base_img = self.generate_poster("", bg_path)
        base_img = base_img.convert('RGBA')

        try:
            main_font = ImageFont.truetype('fonts/simhei.ttf', self.main_size.value())
        except:
            main_font = ImageFont.load_default()

        words = list(text)
        max_line_chars = 10
        lines = []
        current_line = []

        # 定义需要在后面换行的标点符号
        break_after_punctuation = {'，', '。', '；', '！', '？', '」', '》', '》', '、', ',', '.', ';', '!', '?'}

        for i, word in enumerate(words):
            current_line.append(word)
            # 条件1：达到最大字数换行
            # 条件2：遇到标点符号换行（且当前行已经有内容）
            if (len(current_line) >= max_line_chars or
                (word in break_after_punctuation and len(current_line) > 1)):
                lines.append(''.join(current_line))
                current_line = []


        if current_line:
            lines.append(''.join(current_line))

        char_positions = []
        y_pos = 500  # 修改：从600改为500，上移100像素
        line_height = 100

        for line in lines:
            full_width = main_font.getlength(line)
            start_x = 450 - full_width / 2
            x_pos = start_x
            for char in line:
                char_width = main_font.getlength(char)
                char_positions.append((int(x_pos), y_pos, char))
                x_pos += char_width
            y_pos += line_height

        frames = []
        displayed_chars = []

        for i, (x, y, char) in enumerate(char_positions):
            if frames:
                frame_img = Image.fromarray(cv2.cvtColor(frames[-1], cv2.COLOR_BGR2RGBA))
            else:
                frame_img = base_img.copy()

            draw = ImageDraw.Draw(frame_img)
            displayed_chars.append((x, y, char))

            for cx, cy, c in displayed_chars:
                draw.text((cx, cy), c, fill=self.colors['main'], font=main_font)

            frame = cv2.cvtColor(np.array(frame_img), cv2.COLOR_RGBA2BGR)
            frames.extend([frame] * int(fps * char_duration))

        if frames:
            frames.extend([frames[-1]] * int(fps * final_duration))
        else:
            frame = cv2.cvtColor(np.array(base_img), cv2.COLOR_RGBA2BGR)
            frames.append(frame)

        return frames

if __name__ == '__main__':
    app = QApplication(sys.argv)

    if not os.path.exists('fonts/simhei.ttf'):
        try:
            system_fonts = [
                'C:/Windows/Fonts/simhei.ttf',
                'C:/Windows/Fonts/msyh.ttf',
                '/System/Library/Fonts/PingFang.ttf',
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'
            ]
            for font_path in system_fonts:
                if os.path.exists(font_path):
                    os.symlink(font_path, 'fonts/simhei.ttf')
                    break
        except:
            pass

    window = VideoGenerator()
    window.show()
    sys.exit(app.exec_())
